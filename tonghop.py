import openpyxl
import datetime
import os
import sys
import pyexcel
import win32com.client as win32
import shutil
import logging
#ver 3.0 30/01/2019
'''
-convert xls --> xlsx vì package openpyxl không đọc được file xls
-Convert xong chuyển các file xls vào 1 thư mục
-thêm list tên các sheet hợp lệ, nếu user tạo thêm các sheet không nằm trong list này thì ứng dụng bỏ qua ko lấy dữ liệu
-Ghi log tên các file kết quả vào file log\log.txt
'''
#ver 2.2 30/01/2019
'''
-Bắt lỗi file nào không có sheet TT-chung
-Bắt lỗi file nào không đúng cấu trúc (các sheet không khớp như định sẵn)
'''
#ver 2.1: 28/01/2019
'''
-Thêm cột tên file chi tiết vào sheet TT-chung của file tổng hợp để phát hiện sai xót
'''
#ver 2.0: 25/01/2019
'''
- Thêm sheet 10, 11, 12 vào các file tổng hợp -> phòng ngừa thiếu sẽ gây lỗi
- Bắt lỗi nếu có file nào sai cấu trúc
- Thông báo nếu không có file excel nào trong thư mục hiện tại
'''
#ver 1.0: 24/01/2019
now = datetime.datetime.now()
str_now = now.strftime('%Y%m%d%H%M%S')
#tao folder log
if not os.path.exists('log'):
		os.makedirs('log')
file_log = "log\\log_"+str_now+".txt"
logging.basicConfig(filename=file_log, level=logging.DEBUG)
accept_sheets = ['TT-chung','2','3','4','5','6','7','8','9','10','11','12']

#Tìm file xls để convert sang xlsx
file_xls = []
for filename in os.listdir('.'):
	extend_file = filename.split('.')[-1]
	if (extend_file == "xls"):
		file_xls.append(filename)
if (len(file_xls) > 0):
	print(file_xls)
	print("Đã tìm thấy %s file xlsx"%len(file_xls))
	print("Đang convert các file xls sang xlsx")
	dir_path = os.getcwd() #get folder path
	#tao folder cac file xls bi move vao
	if not os.path.exists('moved'):
		os.makedirs('moved')
	for filename in file_xls:
		try:
			fname = dir_path+'\\'+filename #path đầy đủ của file xls
			excel = win32.gencache.EnsureDispatch('Excel.Application')
			wb3 = excel.Workbooks.Open(fname)
			dest_file = 'convert_'+filename+'x' #file ket qua convert
			print('%s --> %s'%(filename, dest_file))
			wb3.SaveAs(dir_path+'\\'+dest_file, FileFormat=51)
			wb3.Close()
			excel.Application.Quit()

			#move xls to folder
			shutil.move(fname, dir_path+'\\moved\\'+filename)

			'''
			filename_detail = filename.split('.')[0]
			dest_file = '%s_convert.xlsx'%filename_detail
			print('%s --> %s'%(filename, dest_file))
			pyexcel.save_book_as(file_name=filename, dest_file_name=dest_file, formatting_info=True)
			'''
		except Exception as e:
			print("Lỗi: file %s"%filename)
			print(e)
			os.system('pause')
			sys.exit("... It's over ...")

#tim file xlsx
file_list = []
for filename in os.listdir('.'):
	extend_file = filename.split('.')[-1]
	if (extend_file == "xlsx"):
		file_list.append(filename)
tong_file_phai_lam = len(file_list)
if tong_file_phai_lam > 0:
	#tao folder chua ket qua:
	if not os.path.exists('ketqua'):
		os.makedirs('ketqua')

	print(file_list)
	#logging.debug(file_list)
	print("Tổng số file %s:" %tong_file_phai_lam)
	print("Loading ..........")

	'''
	sheet = wb['TT-chung']
	row_count = sheet.max_row
	if (row_count > 1):
		row_count = row_count + 1
	'''
	#column_count = sheet.max_column

	#file_list = ['Mau 01.xlsx', 'Mau 11.xlsx', 'Mau 01-2.xlsx']
	#try:
	catalog_file = {}
	for file_name in file_list:
		#wb = openpyxl.load_workbook("tong hop "+file_name)
		#print("Đang đọc file %s"%file_name)
		#logging.debug(file_name)
		wb2 = openpyxl.load_workbook(file_name)
		try:
			sheet = wb2['TT-chung']
			if (sheet['C1'].value in catalog_file):
				catalog_file[sheet['C1'].value].append(file_name)
			else:
				catalog_file[sheet['C1'].value] = [file_name]
		except Exception as e:
			print("Lỗi: file %s không có sheet TT-chung"%file_name)
			print(e)
			os.system('pause')
			sys.exit("... It's over ...")
	#print(catalog_file)
	#print("Loading 2 ..........")
	file_dang_lam = 0
	file_ket_qua = []
	for mau in catalog_file.keys():
		#print("tonghop%s.xlsx" %mau)
		ten_file_tong_hop = "ketqua\\tonghop_mau%s_%s.xlsx" %(mau,str_now)

		#tao file tong hop moi
		wb = openpyxl.Workbook()
		sheet_moi = wb.active
		sheet_moi.title = "TT-chung"
		wb.create_sheet(title="2")
		wb.create_sheet(title="3")
		wb.create_sheet(title="4")
		wb.create_sheet(title="5")
		wb.create_sheet(title="6")
		wb.create_sheet(title="7")
		wb.create_sheet(title="8")
		wb.create_sheet(title="9")
		wb.create_sheet(title="10")
		wb.create_sheet(title="11")
		wb.create_sheet(title="12")
		wb.save(filename = ten_file_tong_hop)
		#end tao file moi

		wb = openpyxl.load_workbook(ten_file_tong_hop) #mo file moi vua tao
		for file_name in catalog_file[mau]:
			file_dang_lam = file_dang_lam + 1
			print("Đang lấy dữ liệu file thứ %s/%s" %(file_dang_lam,tong_file_phai_lam))
			wb2 = openpyxl.load_workbook(file_name)
			all_sheet_wb2 = wb2.sheetnames
			str_mau_khao_sat = ""
			mst = ""

			for sheet_name2 in all_sheet_wb2:
				if (sheet_name2 in accept_sheets):
					try:
						#choose sheet in tonghop.xlsx
						sheet = wb[sheet_name2]
						row_count = sheet.max_row+1
						#if (row_count > 1):
						#	row_count = row_count + 1
						#end choose sheet
						
						sheet2 = wb2[sheet_name2]
						row_count2 = sheet2.max_row
						column_count2 = sheet2.max_column
						for x in range(1,row_count2+1):
							for y in range(1, column_count2+1):
								#ghi gia tri vao file tonghop.xlsx
								if (sheet_name2=="TT-chung"):
									sheet.cell(row = row_count, column = 1).value = file_name #COT 1 ghi ten file
									sheet.cell(row = row_count, column = y+1).value = sheet2.cell(row = x, column = y).value
									if (sheet2.cell(row=x, column=y).value == "MST"):
										mst = sheet2.cell(row=x, column=y+1).value
								else:
									sheet.cell(row = row_count, column = 1).value = mst
									sheet.cell(row = row_count, column = y+1).value = sheet2.cell(row = x, column = y).value
							row_count = row_count + 1
					except Exception as e:
						print("Lỗi: file %s bị lỗi cấu trúc"%file_name)
						print(e)
						os.system('pause')
						sys.exit("... It's over ...")

		#sheet['A1'] = 'hihihi' #ghi gia tri
		file_ket_qua.append(ten_file_tong_hop)
		wb.save(ten_file_tong_hop)
		logging.debug(ten_file_tong_hop)
	print("Đã tổng hợp số liệu thành công. File kết quả là:")
	print(file_ket_qua)
	print("Nhấn phím bất kỳ để thoát...")
	os.system('pause')

	'''
	except Exception as e:
		print("Lỗi: Có file bị lỗi cấu trúc")
		print(e)
		os.system('pause')
	'''
else:
	print("Không có file nào đủ điều kiện để lấy dữ liệu. Nhấn phím bất kỳ để thoát...")
	os.system('pause')